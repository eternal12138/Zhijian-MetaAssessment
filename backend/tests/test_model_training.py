from pathlib import Path
from tempfile import TemporaryDirectory
from types import SimpleNamespace
import unittest

import joblib
import numpy as np
from fastapi import HTTPException

from app.api.model_training import _delete_training_jobs, _managed_artifact_directory, delete_job
from app.config import Settings
from app.models.research import ModelTrainingJob
from app.schemas.model_training import TrainingJobCreate, TrainingJobsDeleteIn, TrainingSuiteCreate
from app.services.model_artifacts import load_model_artifact, sha256_file
from app.services.model_inference import _probe_embedding_features
from app.services.model_training import (
    _TrainingProgressState, _job_config, _job_embedding_config,
    _train_embedding_classifier, _train_tfidf,
)
from app.services.model_training_datasets import (
    build_dataset_template, load_dataset_samples, materialize_dataset, parse_uploaded_dataset,
    get_dataset_metadata,
)
from app.training.baseline_models import (
    build_remote_embedding_linear_svc,
    build_remote_embedding_logistic,
    build_remote_embedding_random_forest,
    build_remote_embedding_xgboost,
    build_remote_embedding_lightgbm,
    build_remote_embedding_catboost,
    build_tfidf_linear_svc,
)
from app.training.hyperparameters import normalize_hyperparameters, public_hyperparameter_catalog
from openpyxl import load_workbook
from io import BytesIO
from scripts.model_training_worker import source_revision


class ModelTrainingLifecycleTests(unittest.TestCase):
    @staticmethod
    def synthetic_samples():
        samples = []
        for group in range(10):
            for label, phrase in enumerate(("检查理解", "调整方法", "评价结果"), start=1):
                samples.append((f"user-{group}", f"{phrase} 第{group}组 标签{label}", label))
        return samples

    def test_training_versions_allow_chinese_but_reject_unsafe_path_characters(self):
        single = TrainingJobCreate(version="元认知模型_v2")
        suite = TrainingSuiteCreate(version_prefix="三分类对比-第一轮")
        self.assertEqual(single.version, "元认知模型_v2")
        self.assertEqual(suite.version_prefix, "三分类对比-第一轮")
        with self.assertRaises(ValueError):
            TrainingJobCreate(version="../模型")
        with self.assertRaises(ValueError):
            TrainingSuiteCreate(version_prefix="模型/第一轮")
        with self.assertRaises(ValueError):
            TrainingSuiteCreate(version_prefix="模型 第一轮")

    def test_batch_delete_ids_must_be_unique_and_non_empty(self):
        self.assertEqual(TrainingJobsDeleteIn(job_ids=[" job-1 "]).job_ids, ["job-1"])
        with self.assertRaises(ValueError):
            TrainingJobsDeleteIn(job_ids=["job-1", "job-1"])
        with self.assertRaises(ValueError):
            TrainingJobsDeleteIn(job_ids=[])

    def test_activation_probe_accepts_numpy_embedding_matrix(self):
        vectors = np.ones((1, 4), dtype=np.float32)
        features = _probe_embedding_features(vectors, 4)
        self.assertEqual(features.shape, (1, 4))
        self.assertTrue(np.isfinite(features).all())

    def test_activation_probe_rejects_wrong_embedding_shape(self):
        with self.assertRaisesRegex(ValueError, "向量形状异常"):
            _probe_embedding_features(np.ones((2, 4), dtype=np.float32), 4)

    def test_custom_comparison_requires_unique_model_selection(self):
        suite = TrainingSuiteCreate(
            version_prefix="自选对比_v1",
            experiment_types=["tfidf_linear_svc", "embedding_logistic"],
        )
        self.assertEqual(
            suite.experiment_types,
            ["tfidf_linear_svc", "embedding_logistic"],
        )
        with self.assertRaisesRegex(ValueError, "至少需要选择两个模型"):
            TrainingSuiteCreate(
                version_prefix="仅一个模型",
                experiment_types=["tfidf_linear_svc"],
            )
        with self.assertRaisesRegex(ValueError, "不能重复选择"):
            TrainingSuiteCreate(
                version_prefix="重复模型",
                experiment_types=["tfidf_linear_svc", "tfidf_linear_svc"],
            )

    def test_four_baseline_builders_are_centralized_and_available(self):
        self.assertEqual(type(build_tfidf_linear_svc()).__name__, "Pipeline")
        self.assertEqual(type(build_remote_embedding_linear_svc()).__name__, "LinearSVC")
        self.assertEqual(type(build_remote_embedding_logistic()).__name__, "LogisticRegression")
        self.assertEqual(type(build_remote_embedding_random_forest()).__name__, "RandomForestClassifier")

    def test_seven_experiments_publish_validated_parameter_contracts(self):
        catalog = public_hyperparameter_catalog()
        self.assertEqual(len(catalog), 7)
        self.assertIn("embedding_xgboost", catalog)
        self.assertIn("embedding_lightgbm", catalog)
        self.assertIn("embedding_catboost", catalog)
        parameters, tuned = normalize_hyperparameters("tfidf_linear_svc", {"C": 2.5})
        self.assertTrue(tuned)
        self.assertEqual(parameters["C"], 2.5)
        self.assertEqual(build_tfidf_linear_svc(hyperparameters=parameters)["classifier"].C, 2.5)

    def test_logistic_class_weight_is_tunable_and_recordable(self):
        catalog = public_hyperparameter_catalog()["embedding_logistic"]
        self.assertEqual(catalog["defaults"]["class_weight"], "balanced")
        self.assertEqual(catalog["parameters"]["class_weight"]["choices"], ["balanced", "none"])

        balanced, tuned = normalize_hyperparameters(
            "embedding_logistic", {"class_weight": "balanced"},
        )
        self.assertTrue(tuned)
        self.assertEqual(build_remote_embedding_logistic(hyperparameters=balanced).class_weight, "balanced")

        unweighted, tuned = normalize_hyperparameters(
            "embedding_logistic", {"class_weight": "none"},
        )
        self.assertTrue(tuned)
        self.assertIsNone(build_remote_embedding_logistic(hyperparameters=unweighted).class_weight)
        with self.assertRaises(ValueError):
            normalize_hyperparameters("embedding_logistic", {"class_weight": "automatic"})

    def test_linear_svc_class_weight_is_tunable_for_both_feature_paths(self):
        for experiment in ("tfidf_linear_svc", "embedding_linear_svc"):
            catalog = public_hyperparameter_catalog()[experiment]
            self.assertEqual(catalog["defaults"]["class_weight"], "balanced")
            self.assertEqual(catalog["parameters"]["class_weight"]["choices"], ["balanced", "none"])

        tfidf_parameters, tuned = normalize_hyperparameters(
            "tfidf_linear_svc", {"class_weight": "none"},
        )
        self.assertTrue(tuned)
        self.assertIsNone(
            build_tfidf_linear_svc(hyperparameters=tfidf_parameters)["classifier"].class_weight
        )

        embedding_parameters, tuned = normalize_hyperparameters(
            "embedding_linear_svc", {"class_weight": "none"},
        )
        self.assertTrue(tuned)
        self.assertIsNone(
            build_remote_embedding_linear_svc(hyperparameters=embedding_parameters).class_weight
        )

    def test_xgboost_complexity_parameters_are_tunable_and_documented(self):
        catalog = public_hyperparameter_catalog()["embedding_xgboost"]
        self.assertIn("n_estimators", catalog["parameters"])
        for name in ("max_depth", "min_child_weight", "gamma", "reg_lambda"):
            self.assertIn(name, catalog["parameters"])
            self.assertTrue(catalog["parameters"][name]["description"].strip())
        parameters, tuned = normalize_hyperparameters(
            "embedding_xgboost", {
                "n_estimators": 200,
                "max_depth": 5,
                "min_child_weight": 2.5,
                "gamma": 0.3,
                "reg_lambda": 1.5,
            },
        )
        self.assertTrue(tuned)
        self.assertEqual(parameters["n_estimators"], 200)
        self.assertEqual(parameters["min_child_weight"], 2.5)
        classifier = build_remote_embedding_xgboost(hyperparameters=parameters)
        self.assertEqual(classifier.estimator.n_estimators, 200)
        self.assertEqual(classifier.estimator.max_depth, 5)
        self.assertEqual(classifier.estimator.min_child_weight, 2.5)
        self.assertEqual(classifier.estimator.gamma, 0.3)
        self.assertEqual(classifier.estimator.reg_lambda, 1.5)

    def test_new_boosting_classifiers_fit_predict_and_reload(self):
        features = np.asarray([
            [label, group / 10, label * group / 10]
            for group in range(6) for label in (1, 2, 3)
        ], dtype=np.float32)
        labels = np.asarray([label for _group in range(6) for label in (1, 2, 3)])
        builders = [
            (build_remote_embedding_xgboost, {"n_estimators": 10, "max_depth": 3}),
            (build_remote_embedding_lightgbm, {"n_estimators": 10, "num_leaves": 7}),
            (build_remote_embedding_catboost, {"iterations": 10, "depth": 3}),
        ]
        with TemporaryDirectory() as directory:
            for index, (builder, parameters) in enumerate(builders):
                classifier = builder(hyperparameters=parameters)
                classifier.fit(features, labels)
                self.assertTrue(set(np.asarray(classifier.predict(features)).reshape(-1).tolist()).issubset({1, 2, 3}))
                path = Path(directory) / f"boosting-{index}.joblib"
                joblib.dump(classifier, path)
                restored = joblib.load(path)
                self.assertEqual(restored.predict(features[:1]).shape, (1,))

    def test_hyperparameter_contract_rejects_unknown_or_out_of_range_values(self):
        with self.assertRaises(ValueError):
            normalize_hyperparameters("embedding_xgboost", {"unknown": 1})
        with self.assertRaises(ValueError):
            normalize_hyperparameters("embedding_catboost", {"depth": 99})

    def test_job_uses_frozen_embedding_configuration(self):
        job = SimpleNamespace(config_snapshot={
            "embedding_base_url": "https://frozen.example/v1",
            "embedding_model": "frozen-model",
            "dimensions": 768,
            "batch_size": 8,
            "timeout_seconds": 45,
        })
        runtime = SimpleNamespace(
            QWEN_EMBEDDING_BASE_URL="https://changed.example/v1",
            QWEN_EMBEDDING_MODEL="changed-model",
            QWEN_EMBEDDING_DIMENSIONS=1024,
            QWEN_EMBEDDING_BATCH_SIZE=20,
            QWEN_EMBEDDING_TIMEOUT_SECONDS=60,
            QWEN_EMBEDDING_API_KEY="runtime-secret",
        )
        config = _job_config(job, runtime)
        self.assertEqual(config["base_url"], "https://frozen.example/v1")
        self.assertEqual(config["model"], "frozen-model")
        self.assertEqual(config["dimensions"], 768)
        self.assertEqual(config["api_key"], "runtime-secret")

    def test_tfidf_job_does_not_parse_optional_embedding_configuration(self):
        job = SimpleNamespace(config_snapshot={"feature": "tfidf", "max_retries": None})
        runtime = SimpleNamespace()
        self.assertIsNone(_job_embedding_config(job, runtime))

    def test_embedding_optional_limits_use_safe_defaults(self):
        job = SimpleNamespace(config_snapshot={
            "feature": "remote_embedding",
            "embedding_base_url": "https://embedding.example/v1",
            "embedding_model": "embed-v1",
            "dimensions": 1024,
            "batch_size": None,
            "timeout_seconds": None,
            "max_retries": None,
        })
        runtime = SimpleNamespace(
            EMBEDDING_PROVIDER="openai_compatible",
            EMBEDDING_API_KEY="secret",
            EMBEDDING_BATCH_SIZE=None,
            EMBEDDING_TIMEOUT=None,
            EMBEDDING_MAX_RETRIES=None,
            EMBEDDING_NORMALIZED=True,
            EMBEDDING_INSTRUCTION="",
        )
        config = _job_embedding_config(job, runtime)
        self.assertIsNotNone(config)
        self.assertEqual(config["batch_size"], 32)
        self.assertEqual(config["timeout"], 60.0)
        self.assertEqual(config["max_retries"], 4)

    def test_embedding_missing_dimension_has_actionable_error(self):
        job = SimpleNamespace(config_snapshot={
            "feature": "remote_embedding",
            "embedding_base_url": "https://embedding.example/v1",
            "embedding_model": "embed-v1",
        })
        runtime = SimpleNamespace(EMBEDDING_MODEL="", QWEN_EMBEDDING_DIMENSIONS=None)
        with self.assertRaisesRegex(ValueError, "向量维度必须是整数"):
            _job_embedding_config(job, runtime)

    def test_artifact_integrity_is_required_before_activation(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            artifact = root / "v1" / "model.joblib"
            artifact.parent.mkdir()
            joblib.dump({
                "scaler": "scaler", "classifier": "classifier",
                "embedding_model": "qwen", "dimensions": 1024,
            }, artifact)
            job = ModelTrainingJob(
                version="v1", requested_by="admin", artifact_path=str(artifact),
                artifact_sha256=sha256_file(artifact), status="completed",
            )
            settings = Settings(MODEL_TRAINING_DIR=str(root), _env_file=None)
            loaded = load_model_artifact(job, settings)
            self.assertEqual(loaded["dimensions"], 1024)
            job.artifact_sha256 = "0" * 64
            with self.assertRaisesRegex(ValueError, "完整性"):
                load_model_artifact(job, settings)

    def test_training_record_cleanup_is_limited_to_managed_version_directory(self):
        with TemporaryDirectory() as directory:
            root = Path(directory).resolve()
            managed = root / "v1" / "model.joblib"
            managed.parent.mkdir()
            managed.touch()
            job = ModelTrainingJob(
                version="v1", requested_by="admin", artifact_path=str(managed),
            )
            self.assertEqual(_managed_artifact_directory(job, root), root / "v1")

            outside = root.parent / "outside-model.joblib"
            job.artifact_path = str(outside)
            job.version = "../outside"
            self.assertIsNone(_managed_artifact_directory(job, root))

    def test_tfidf_uses_group_folds_and_returns_complete_metrics(self):
        samples = self.synthetic_samples()
        labels = np.asarray([item[2] for item in samples], dtype=np.int64)
        groups = np.asarray([item[0] for item in samples])
        _, pipeline, metrics = _train_tfidf(
            samples, labels, groups, {"C": 0.75, "max_iter": 3000},
        )
        self.assertEqual(len(metrics["folds"]), 5)
        self.assertEqual(len(metrics["confusion_matrix"]), 3)
        self.assertEqual(set(metrics["per_class"]), {"1", "2", "3"})
        self.assertIn("macro_specificity", metrics)
        self.assertIn("weighted_precision", metrics)
        self.assertIn("weighted_recall", metrics)
        self.assertEqual(len(metrics["_oof_predictions"]), len(samples))
        self.assertTrue(all("specificity" in item for item in metrics["per_class"].values()))
        self.assertTrue(all(fold["subject_disjoint_verified"] for fold in metrics["folds"]))
        self.assertTrue(all(isinstance(fold.get("macro_auc_ovr"), float) and 0.0 <= fold["macro_auc_ovr"] <= 1.0 for fold in metrics["folds"]))
        self.assertTrue(all(set(fold.get("per_class_auc", {}).keys()) == {"1", "2", "3"} for fold in metrics["folds"]))
        self.assertEqual(set(metrics["roc_curves"]), {"macro", "1", "2", "3"})
        self.assertIn("macro_auc_ovr", metrics)
        self.assertNotIn("cross_entropy", metrics)
        self.assertEqual(metrics["roc_evaluation"]["source"], "cross_validated_out_of_fold")
        self.assertEqual(metrics["roc_evaluation"]["score_type"], "decision_function")
        self.assertFalse(metrics["roc_evaluation"]["external_holdout"])
        self.assertEqual(sum(fold["sample_count"] for fold in metrics["folds"]), len(samples))
        self.assertTrue(all(fold["train_sample_count"] + fold["sample_count"] == len(samples) for fold in metrics["folds"]))
        self.assertTrue(all(set(fold["test_label_distribution"]) == {"1", "2", "3"} for fold in metrics["folds"]))
        self.assertEqual(len(pipeline.predict(["我检查自己的理解"])), 1)

    def test_training_reports_each_fold_and_final_refit(self):
        samples = self.synthetic_samples()
        labels = np.asarray([item[2] for item in samples], dtype=np.int64)
        groups = np.asarray([item[0] for item in samples])
        events = []
        _train_tfidf(
            samples, labels, groups, {"C": 0.75, "max_iter": 3000},
            lambda event, fold, total: events.append((event, fold, total)),
        )
        self.assertEqual(
            [item for item in events if item[0] == "fold_started"],
            [("fold_started", fold, 5) for fold in range(1, 6)],
        )
        self.assertEqual(events[-2:], [("refit_started", 5, 5), ("refit_completed", 5, 5)])

    def test_training_progress_state_exposes_fold_progress_and_eta(self):
        state = _TrainingProgressState(60)
        state.callback("fold_started", 1, 5)
        state._unit_started_at -= 2
        state.callback("fold_completed", 1, 5)
        state.callback("fold_started", 2, 5)
        snapshot = state.snapshot()
        self.assertEqual(snapshot["stage"], "training_fold_2")
        self.assertEqual(snapshot["current_fold"], 2)
        self.assertEqual(snapshot["total_folds"], 5)
        self.assertGreater(snapshot["progress"], 60)
        self.assertIsNotNone(snapshot["estimated_remaining_seconds"])

    def test_six_embedding_classifiers_return_complete_metrics(self):
        samples = self.synthetic_samples()
        labels = np.asarray([item[2] for item in samples], dtype=np.int64)
        groups = np.asarray([item[0] for item in samples])
        features = np.asarray([
            [label == index for index in (1, 2, 3)] + [group / 10]
            for group in range(10) for label in (1, 2, 3)
        ], dtype=np.float32)
        parameter_sets = {
            "linear_svc": {"C": 0.75, "max_iter": 3000},
            "logistic": {"C": 0.75, "max_iter": 1000},
            "random_forest": {"n_estimators": 20, "max_depth": 4, "min_samples_leaf": 1, "max_features": "sqrt"},
            "xgboost": {"n_estimators": 10, "max_depth": 3, "learning_rate": .1, "subsample": .9, "colsample_bytree": .9, "reg_alpha": 0, "reg_lambda": 1},
            "lightgbm": {"n_estimators": 10, "num_leaves": 7, "max_depth": 4, "learning_rate": .1, "min_child_samples": 2, "subsample": .9, "colsample_bytree": .9, "reg_alpha": 0, "reg_lambda": 0},
            "catboost": {"iterations": 10, "depth": 3, "learning_rate": .1, "l2_leaf_reg": 3, "random_strength": 1},
        }
        for classifier_type, parameters in parameter_sets.items():
            with self.subTest(classifier=classifier_type):
                _, classifier, metrics = _train_embedding_classifier(
                    features, labels, groups, classifier_type, parameters,
                )
                self.assertEqual(len(metrics["folds"]), 5)
                self.assertIn("weighted_f1", metrics)
                self.assertIn("macro_specificity", metrics)
                self.assertEqual(set(metrics["roc_curves"]), {"macro", "1", "2", "3"})
                self.assertTrue(all(
                    len(curve["fpr"]) == len(curve["tpr"])
                    for curve in metrics["roc_curves"].values()
                ))
                self.assertEqual(
                    metrics["roc_evaluation"]["score_type"],
                    "decision_function" if classifier_type == "linear_svc" else "predict_proba",
                )
                self.assertTrue(all("train_macro_f1" in fold for fold in metrics["folds"]))
                self.assertTrue(all(isinstance(fold.get("macro_auc_ovr"), float) and 0.0 <= fold["macro_auc_ovr"] <= 1.0 for fold in metrics["folds"]))
                self.assertEqual(len(classifier.predict(features[:1])), 1)
                if classifier_type == "linear_svc":
                    self.assertNotIn("cross_entropy", metrics)
                else:
                    self.assertIn("cross_entropy", metrics)

    def test_uploaded_dataset_is_validated_and_materialized_immutably(self):
        rows = ["participant_id,clean_text,label"]
        for group in range(10):
            for label in (0, 1, 2, 3):
                rows.append(f"user-{group},第{group}名被试的第{label}类文本,{label}")
        samples = parse_uploaded_dataset("gold.csv", "\n".join(rows).encode("utf-8"))
        with TemporaryDirectory() as directory:
            metadata = materialize_dataset(
                Path(directory), samples, source="uploaded", name="外部金标准 v1",
                original_filename="gold.csv", created_by="admin",
            )
            loaded = load_dataset_samples(Path(directory), metadata["id"])
        self.assertEqual(samples, loaded)
        self.assertEqual(metadata["sample_count"], 40)
        self.assertEqual(metadata["training_sample_count"], 30)
        self.assertEqual(metadata["excluded_non_metacognitive_count"], 10)
        self.assertEqual(metadata["participant_count"], 10)
        self.assertEqual(metadata["label_distribution"], {"0": 10, "1": 10, "2": 10, "3": 10})

    def test_uploaded_dataset_accepts_clean_text_and_label_only(self):
        rows = ["clean_text,label"]
        for index in range(10):
            for label in (1, 2, 3):
                rows.append(f"第{index}条第{label}类清洗后文本,{label}")
        samples = parse_uploaded_dataset("minimal.csv", "\n".join(rows).encode("utf-8"))
        self.assertEqual(len(samples), 30)
        self.assertTrue(all(participant == "" for participant, _, _ in samples))
        with TemporaryDirectory() as directory:
            metadata = materialize_dataset(
                Path(directory), samples, source="uploaded", name="两列训练数据",
                original_filename="minimal.csv", created_by="admin",
            )
            loaded = load_dataset_samples(Path(directory), metadata["id"])
        self.assertEqual(samples, loaded)
        self.assertEqual(metadata["participant_count"], 0)
        self.assertFalse(metadata["has_participant_ids"])
        self.assertEqual(metadata["split_strategy"], "sentence_stratified_5fold")

    def test_uploaded_dataset_accepts_exported_training_column_names(self):
        rows = ["cleaned_text,label_train"]
        for index in range(10):
            for label in (1, 2, 3):
                rows.append(f"第{index}条第{label}类清洗文本,{label}")
        samples = parse_uploaded_dataset("exported.csv", "\n".join(rows).encode("utf-8"))
        self.assertEqual(len(samples), 30)
        self.assertEqual({item[2] for item in samples}, {1, 2, 3})

    def test_conflicting_text_error_contains_actionable_preview(self):
        samples = [
            ("", f"普通样本{label}-{index}", label)
            for label in (1, 2, 3) for index in range(10)
        ]
        samples.extend([("", "计算方差这个比较复杂。", 1), ("", "计算方差这个比较复杂。", 3)])
        with self.assertRaisesRegex(ValueError, "计算方差这个比较复杂"):
            materialize_dataset(
                Path("unused"), samples, source="uploaded", name="冲突数据",
                original_filename="conflict.csv", created_by="admin",
            )

    def test_tfidf_supports_sentence_stratified_folds_without_participant_ids(self):
        samples = [
            ("", f"第{index}条第{label}类文本", label)
            for index in range(10) for label in (1, 2, 3)
        ]
        labels = np.asarray([item[2] for item in samples], dtype=np.int64)
        groups = np.asarray([item[0] for item in samples])
        _, pipeline, metrics = _train_tfidf(samples, labels, groups)
        self.assertEqual(len(metrics["folds"]), 5)
        self.assertEqual(len(pipeline.predict(["检查当前结果"])), 1)

    def test_legacy_dataset_metadata_infers_subject_grouped_split(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            dataset_id = "11111111-1111-1111-1111-111111111111"
            target = root / "datasets"
            target.mkdir()
            (target / f"{dataset_id}.json").write_text(
                '{"id":"' + dataset_id + '","participant_count":8}', encoding="utf-8"
            )
            metadata = get_dataset_metadata(root, dataset_id)
        self.assertTrue(metadata["has_participant_ids"])
        self.assertEqual(metadata["split_strategy"], "subject_grouped_stratified_5fold")

    def test_uploaded_dataset_rejects_unconfirmed_taxonomy(self):
        content = "participant_id,text,label\nuser-1,测试文本,planning\n".encode("utf-8")
        with self.assertRaisesRegex(ValueError, "未支持的标签"):
            parse_uploaded_dataset("gold.csv", content)

    def test_download_template_contains_data_and_instructions(self):
        workbook = load_workbook(BytesIO(build_dataset_template()), read_only=True, data_only=True)
        try:
            self.assertEqual(workbook.sheetnames, ["训练数据模板", "填写说明"])
            headers = [cell.value for cell in workbook["训练数据模板"][1]]
            self.assertEqual(headers, ["clean_text", "label"])
            self.assertIn("请先删除", str(workbook["填写说明"]["B1"].value))
        finally:
            workbook.close()

    def test_worker_source_revision_changes_with_source(self):
        with TemporaryDirectory() as directory:
            source = Path(directory) / "worker-source.py"
            source.write_text("VERSION = 1", encoding="utf-8")
            first = source_revision((source,))
            source.write_text("VERSION = 2", encoding="utf-8")
            second = source_revision((source,))
        self.assertNotEqual(first, second)


class _TrainingDeleteDb:
    def __init__(self, job):
        self.job = job
        self.added = []
        self.deleted = []
        self.committed = False

    async def scalar(self, _statement):
        return self.job

    def add(self, item):
        self.added.append(item)

    async def delete(self, item):
        self.deleted.append(item)

    async def commit(self):
        self.committed = True


class ModelTrainingDeleteTests(unittest.IsolatedAsyncioTestCase):
    async def test_terminal_inactive_record_can_be_deleted(self):
        job = SimpleNamespace(
            id="delete-job", version="deleted-v1", status="failed",
            is_active=False, artifact_path=None,
        )
        db = _TrainingDeleteDb(job)

        result = await delete_job(
            job.id, db=db, user=SimpleNamespace(id="admin-1")
        )

        self.assertEqual(result["status"], "deleted")
        self.assertEqual(db.deleted, [job])
        self.assertTrue(db.committed)

    async def test_active_record_cannot_be_deleted(self):
        job = SimpleNamespace(
            id="active-job", version="active-v1", status="completed",
            is_active=True, artifact_path=None,
        )
        with self.assertRaises(HTTPException) as context:
            await delete_job(
                job.id,
                db=_TrainingDeleteDb(job),
                user=SimpleNamespace(id="admin-1"),
            )
        self.assertEqual(context.exception.status_code, 409)

    async def test_batch_delete_validates_every_record_before_deleting(self):
        terminal = SimpleNamespace(
            id="terminal-job", version="terminal-v1", status="failed",
            is_active=False, artifact_path=None,
        )
        running = SimpleNamespace(
            id="running-job", version="running-v1", status="running",
            is_active=False, artifact_path=None,
        )
        db = _TrainingDeleteDb(None)
        with self.assertRaises(HTTPException) as context:
            await _delete_training_jobs(
                [terminal, running], db, SimpleNamespace(id="admin-1"),
            )
        self.assertEqual(context.exception.status_code, 409)
        self.assertEqual(db.deleted, [])
        self.assertFalse(db.committed)

    async def test_batch_delete_removes_all_terminal_records(self):
        jobs = [
            SimpleNamespace(
                id=f"delete-{index}", version=f"deleted-v{index}", status="completed",
                is_active=False, artifact_path=None,
            )
            for index in range(2)
        ]
        db = _TrainingDeleteDb(None)
        result = await _delete_training_jobs(jobs, db, SimpleNamespace(id="admin-1"))
        self.assertEqual([item["job_id"] for item in result], ["delete-0", "delete-1"])
        self.assertEqual(db.deleted, jobs)
        self.assertTrue(db.committed)


if __name__ == "__main__":
    unittest.main()
