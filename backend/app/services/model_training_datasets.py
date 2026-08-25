from __future__ import annotations

import csv
import hashlib
import io
import json
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


LABEL_ALIASES = {
    "0": 0,
    "non_metacognitive": 0,
    "non-meta": 0,
    "non_meta": 0,
    "非元认知": 0,
    "不确定": 0,
    "不确定/不算元认知": 0,
    "1": 1,
    "monitoring": 1,
    "monitor": 1,
    "监控": 1,
    "2": 2,
    "regulation": 2,
    "control": 2,
    "control_regulation": 2,
    "调控": 2,
    "控制": 2,
    "控制/调控": 2,
    "3": 3,
    "evaluation": 3,
    "评估": 3,
}

PARTICIPANT_COLUMNS = ("participant_id", "user_id", "被试编号", "被试id", "用户id", "账号")
TEXT_COLUMNS = (
    "clean_text", "cleaned_text", "text", "清洗后文本", "清洗文本",
    "人工校对文本", "ai筛选后且人工校对的文本", "ai筛选后的转录文本", "文本",
)
LABEL_COLUMNS = (
    "label", "label_train", "expert_label", "final_label", "专家标签", "最终标签", "标签",
)
MAX_DATASET_ROWS = 100_000


def build_dataset_template() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "训练数据模板"
    headers = ["clean_text", "label"]
    sheet.append(headers)
    examples = [
        ["我需要检查刚才的计算是否正确。", 1],
        ["这个方法不合适，我换一种方法。", 2],
        ["这个答案总体上比较合理。", 3],
    ]
    for row in examples:
        sheet.append(row)
    header_fill = PatternFill("solid", fgColor="4F46E5")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in sheet.iter_rows(min_row=2, max_row=4):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="FFF7D6")
    widths = [64, 14]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:B4"
    validation = DataValidation(type="list", formula1='"1,2,3"', allow_blank=False)
    validation.error = "训练标签仅允许填写 1、2、3"
    validation.errorTitle = "标签无效"
    sheet.add_data_validation(validation)
    validation.add("B2:B100001")

    guide = workbook.create_sheet("填写说明")
    guide_rows = [
        ("使用前须知", "请先删除训练数据模板中的黄色 SAMPLE 示例行，再粘贴真实训练数据。"),
        ("必填字段", "只需要 clean_text、label 两列。系统仍兼容旧版含 participant_id 等追溯字段的数据。"),
        ("clean_text", "用于模型训练的人工确认文本，不应放入未经复核的 AI 初步标签。"),
        ("label=1", "监控"),
        ("label=2", "调控"),
        ("label=3", "评估"),
        ("训练范围", "仅训练标签1监控、2调控、3评估；标签0非元认知不进入新模型训练。"),
        ("最低要求", "至少 30 条有效元认知样本，1/2/3 三类均有数据，且每类至少 5条。"),
        ("评估方式", "仅上传文本和标签时采用句子级分层五折；如额外提供 participant_id，则采用更严格的被试级分组五折。"),
        ("数据冲突", "完全相同的文本不能对应多个不同标签；发现冲突时需先完成仲裁。"),
        ("隐私要求", "不要填写姓名、电话、微信号等直接身份信息。"),
    ]
    for row in guide_rows:
        guide.append(row)
    guide.column_dimensions["A"].width = 22
    guide.column_dimensions["B"].width = 92
    guide.freeze_panes = "A2"
    for cell in guide[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    for row in guide.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _normalized_header(value: object) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def _pick_column(headers: list[str], aliases: tuple[str, ...], label: str) -> str:
    lookup = {_normalized_header(item): item for item in headers}
    for alias in aliases:
        if _normalized_header(alias) in lookup:
            return lookup[_normalized_header(alias)]
    raise ValueError(f"训练数据缺少{label}列；可使用：{', '.join(aliases)}")


def _parse_label(value: object) -> int:
    if isinstance(value, bool):
        raise ValueError(f"标签值无效：{value}")
    if isinstance(value, (int, float)) and float(value).is_integer():
        key = str(int(value))
    else:
        key = str(value or "").strip().lower()
    if key not in LABEL_ALIASES:
        raise ValueError(f"发现未支持的标签：{value}；允许 0非元认知、1监控、2调控、3评估，其中0不参与训练")
    return LABEL_ALIASES[key]


def _rows_from_csv(content: bytes) -> list[dict[str, object]]:
    decoded = None
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            decoded = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if decoded is None:
        raise ValueError("CSV 编码无法识别，请另存为 UTF-8 CSV 后重试")
    return list(csv.DictReader(io.StringIO(decoded)))


def _rows_from_xlsx(content: bytes) -> list[dict[str, object]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(iterator, ())]
        if not any(headers):
            return []
        rows: list[dict[str, object]] = []
        for values in iterator:
            rows.append({headers[index]: value for index, value in enumerate(values[:len(headers)])})
            if len(rows) > MAX_DATASET_ROWS:
                raise ValueError(f"训练数据不能超过 {MAX_DATASET_ROWS} 行")
        return rows
    finally:
        workbook.close()


def parse_uploaded_dataset(filename: str, content: bytes) -> list[tuple[str, str, int]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        rows = _rows_from_csv(content)
    elif suffix == ".xlsx":
        rows = _rows_from_xlsx(content)
    else:
        raise ValueError("仅支持 .csv 或 .xlsx 训练数据")
    if not rows:
        raise ValueError("训练数据为空")
    if len(rows) > MAX_DATASET_ROWS:
        raise ValueError(f"训练数据不能超过 {MAX_DATASET_ROWS} 行")
    headers = list(rows[0].keys())
    try:
        participant_column = _pick_column(headers, PARTICIPANT_COLUMNS, "被试标识")
    except ValueError:
        participant_column = None
    text_column = _pick_column(headers, TEXT_COLUMNS, "文本")
    label_column = _pick_column(headers, LABEL_COLUMNS, "标签")
    parsed: list[tuple[str, str, int]] = []
    empty_rows = 0
    for row_number, row in enumerate(rows, start=2):
        participant = str(row.get(participant_column) or "").strip() if participant_column else ""
        text = str(row.get(text_column) or "").strip()
        raw_label = row.get(label_column)
        if not participant and not text and (raw_label is None or str(raw_label).strip() == ""):
            continue
        if not text:
            empty_rows += 1
            continue
        try:
            label = _parse_label(raw_label)
        except ValueError as error:
            raise ValueError(f"第 {row_number} 行：{error}") from error
        parsed.append((participant, text, label))
    if empty_rows:
        raise ValueError(f"发现 {empty_rows} 行缺少清洗后文本，请补全后重新上传")
    return validate_samples(parsed)


def validate_samples(samples: Iterable[tuple[str, str, int]]) -> list[tuple[str, str, int]]:
    unique = sorted(set(samples), key=lambda row: (row[0], row[2], row[1]))
    training_rows = [item for item in unique if item[2] in {1, 2, 3}]
    errors: list[str] = []
    labels_by_text: dict[tuple[str, str], set[int]] = {}
    for participant, text, label in training_rows:
        labels_by_text.setdefault((participant, text), set()).add(label)
    conflicts = [key for key, labels in labels_by_text.items() if len(labels) > 1]
    if conflicts:
        preview = "；".join(
            f"“{text[:40]}{'…' if len(text) > 40 else ''}”"
            for _, text in conflicts[:3]
        )
        errors.append(
            f"检测到 {len(conflicts)} 条相同文本对应多个标签，请先完成仲裁。冲突示例：{preview}"
        )
    if len(training_rows) < 30:
        errors.append("标签1/2/3的有效训练样本少于 30 条")
    labels = {item[2] for item in training_rows}
    if labels != {1, 2, 3}:
        missing = sorted({1, 2, 3} - labels)
        errors.append(f"训练数据必须包含标签1监控、2调控、3评估，当前缺少：{missing}")
    distribution = {label: sum(1 for item in training_rows if item[2] == label) for label in (1, 2, 3)}
    sparse = [label for label, count in distribution.items() if label in labels and count < 5]
    if sparse:
        errors.append(f"每类至少需要 5 条样本才能进行五折评估，当前不足的标签：{sparse}")
    provided_participants = [bool(item[0]) for item in training_rows]
    if any(provided_participants) and not all(provided_participants):
        errors.append("participant_id 列只能全部填写或整列删除；仅使用清洗后文本和标签时请删除该列")
    participants = {item[0] for item in training_rows if item[0]}
    if participants and len(participants) < 5:
        errors.append("提供 participant_id 时至少需要 5 名不同被试；也可以删除该列后采用句子级分层五折")
    if errors:
        raise ValueError("；".join(errors))
    return unique


def _dataset_directory(root: Path) -> Path:
    target = root / "datasets"
    target.mkdir(parents=True, exist_ok=True)
    return target


def materialize_dataset(
    root: Path, samples: Iterable[tuple[str, str, int]], *, source: str,
    name: str, original_filename: str | None, created_by: str,
) -> dict:
    validated = validate_samples(samples)
    canonical = json.dumps(validated, ensure_ascii=False, separators=(",", ":"))
    fingerprint = hashlib.sha256(canonical.encode("utf-8")).hexdigest()
    dataset_id = str(uuid.uuid4())
    directory = _dataset_directory(root)
    csv_path = directory / f"{dataset_id}.csv"
    metadata_path = directory / f"{dataset_id}.json"
    with csv_path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.writer(stream)
        writer.writerow(["participant_id", "clean_text", "label"])
        writer.writerows(validated)
    distribution = {str(label): sum(1 for item in validated if item[2] == label) for label in range(4)}
    training_distribution = {
        str(label): sum(1 for item in validated if item[2] == label) for label in (1, 2, 3)
    }
    metadata = {
        "id": dataset_id,
        "name": name.strip()[:100] or "未命名训练数据",
        "source": source,
        "original_filename": original_filename,
        "sample_count": len(validated),
        "training_sample_count": sum(training_distribution.values()),
        "excluded_non_metacognitive_count": distribution["0"],
        "participant_count": len({item[0] for item in validated if item[0]}),
        "has_participant_ids": all(bool(item[0]) for item in validated),
        "split_strategy": (
            "subject_grouped_stratified_5fold"
            if all(bool(item[0]) for item in validated)
            else "sentence_stratified_5fold"
        ),
        "label_distribution": distribution,
        "training_label_distribution": training_distribution,
        "training_labels": [1, 2, 3],
        "fingerprint": fingerprint,
        "created_by": created_by,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    metadata_path.write_text(json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8")
    return metadata


def get_dataset_metadata(root: Path, dataset_id: str) -> dict:
    try:
        uuid.UUID(dataset_id)
    except ValueError as error:
        raise ValueError("训练数据版本标识无效") from error
    path = _dataset_directory(root) / f"{dataset_id}.json"
    if not path.is_file():
        raise ValueError("训练数据版本不存在或文件已丢失")
    return _normalize_metadata(json.loads(path.read_text(encoding="utf-8")))


def _normalize_metadata(metadata: dict) -> dict:
    """Keep dataset metadata created before the explicit split strategy compatible."""
    normalized = dict(metadata)
    if "has_participant_ids" not in normalized:
        normalized["has_participant_ids"] = int(normalized.get("participant_count") or 0) > 0
    if "split_strategy" not in normalized:
        normalized["split_strategy"] = (
            "subject_grouped_stratified_5fold"
            if normalized["has_participant_ids"]
            else "sentence_stratified_5fold"
        )
    if "training_sample_count" not in normalized:
        distribution = normalized.get("label_distribution") or {}
        normalized["training_sample_count"] = sum(
            int(distribution.get(str(label), 0)) for label in (1, 2, 3)
        )
    normalized.setdefault(
        "excluded_non_metacognitive_count",
        int((normalized.get("label_distribution") or {}).get("0", 0)),
    )
    normalized.setdefault("training_label_distribution", {
        str(label): int((normalized.get("label_distribution") or {}).get(str(label), 0))
        for label in (1, 2, 3)
    })
    normalized.setdefault("training_labels", [1, 2, 3])
    return normalized


def list_uploaded_datasets(root: Path) -> list[dict]:
    rows: list[dict] = []
    for path in _dataset_directory(root).glob("*.json"):
        try:
            metadata = _normalize_metadata(json.loads(path.read_text(encoding="utf-8")))
        except (OSError, ValueError, TypeError):
            continue
        if metadata.get("source") == "uploaded":
            rows.append(metadata)
    return sorted(rows, key=lambda item: str(item.get("created_at") or ""), reverse=True)


def load_dataset_samples(root: Path, dataset_id: str) -> list[tuple[str, str, int]]:
    metadata = get_dataset_metadata(root, dataset_id)
    path = _dataset_directory(root) / f"{dataset_id}.csv"
    if not path.is_file():
        raise ValueError("训练数据文件已丢失")
    with path.open("r", encoding="utf-8-sig", newline="") as stream:
        rows = list(csv.DictReader(stream))
    samples = [
        (row.get("participant_id", ""), row.get("clean_text") or row.get("text") or "", int(row["label"]))
        for row in rows
    ]
    validated = validate_samples(samples)
    canonical = json.dumps(validated, ensure_ascii=False, separators=(",", ":"))
    fingerprint = hashlib.sha256(canonical.encode("utf-8")).hexdigest()
    if fingerprint != metadata.get("fingerprint"):
        raise ValueError("训练数据完整性校验失败，文件内容可能已改变")
    return validated
